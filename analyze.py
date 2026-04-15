"""
MaRisk-Vergleichsversion → Excel-Änderungsanalyse, aggregiert pro Textziffer.

Hauptskript für die Auswertung. Die PDF-Parser-Helfer
(load_page, group_paragraphs, split_footnote_from_tz, …) liegen in
marisk_parser.py. Dieses Skript baut darauf die Tz-Zeilen, fasst pro
Textziffer zusammen und schreibt die Excel-Datei.

Spaltenschema:
  A Textziffer (neu)       B alte Referenz     C Normtext
  D Erläuterung            E Änderungsart Norm F Änderungsart Erl.
  G Verschiebung           H Unsicher          I Anmerkungen
"""
import re
import fitz
from difflib import SequenceMatcher
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font

from marisk_parser import (
    load_page, group_paragraphs, plain_text, segs_to_rich,
    classify_entry, SECTION_RE, _postprocess_xlsx,
    split_footnote_from_tz, diff_summary,
)

PDF = "dl_kon_02_2026_rs_marisk-novelle_vergleichsversion.pdf"
OUT = "MaRisk_Aenderungsanalyse_pro_Textziffer.xlsx"


def build_tz_rows(paragraphs):
    """Collapse paragraphs into one row per (section, Tz).

    Yields dicts with keys:
      label, section, norm_segs, expl_segs, notes, kind
    `kind` is either 'section_header' (pure heading row) or 'tz'.
    """
    current_section = ""
    current_section_old = ""   # old section code if renamed, else ""
    current_tz = None
    current_tz_old = None      # old Tz number if renumbered, else None
    last_new_tz = None         # integer counter of the last real "new" Tz number;
                               # drives the expected-next split context but is
                               # NOT touched by purely deleted old Tz rows.
    current_tz_is_alt = False  # True when the current row is a purely deleted
                               # old Tz (label "alt Tz. N"); suppresses collisions
                               # with the new numbering.
    tz_norm = []
    tz_expl = []
    tz_notes = []
    rows = []
    drop_next_footnote_body = False
    renames = []               # list of (old_code, new_code) for the Umbenennungen sheet

    def flush_tz():
        if current_tz is None and not tz_norm and not tz_expl:
            return
        if current_tz_is_alt:
            # Purely deleted old Tz – label with "alt Tz. N" so it does
            # not collide with the new Tz of the same number.
            label = (f"{current_section} alt Tz. {current_tz}"
                     if current_tz else current_section)
            old_ref = ""
        else:
            label = (f"{current_section} Tz. {current_tz}"
                     if current_tz else current_section)
            old_section_val = current_section_old or current_section
            old_tz_val      = current_tz_old or current_tz
            if current_section_old or current_tz_old:
                old_ref = (f"{old_section_val} Tz. {old_tz_val}"
                           if old_tz_val else old_section_val)
            else:
                old_ref = ""
        rows.append({
            "kind": "tz",
            "label": label,
            "old_ref": old_ref,
            "section": current_section,
            "tz": current_tz,
            "norm_segs": list(tz_norm),
            "expl_segs": list(tz_expl),
            "notes": list(tz_notes),
        })

    for p in paragraphs:
        txt = plain_text(p["segs"]).strip()
        if not txt:
            continue
        kind = p["kind"]
        col  = p["col"]

        if kind in ("section", "subsection"):
            flush_tz()
            tz_norm.clear(); tz_expl.clear(); tz_notes.clear()
            current_tz = None
            current_tz_old = None
            current_tz_is_alt = False
            last_new_tz = None
            drop_next_footnote_body = False
            # Extract new and old section codes.
            kept = plain_text(p["segs"],
                              fmt_filter={"unchanged", "added", "moved_to"}).strip()
            old  = plain_text(p["segs"],
                              fmt_filter={"unchanged", "deleted", "moved_from"}).strip()
            m_new = SECTION_RE.match(kept or txt)
            m_old = SECTION_RE.match(old)
            if m_new:
                current_section = m_new.group(0).replace("  ", " ").strip()
            current_section_old = ""
            if m_new and m_old:
                new_code = m_new.group(0).replace("  ", " ").strip()
                old_code = m_old.group(0).replace("  ", " ").strip()
                if new_code != old_code:
                    current_section_old = old_code
                    renames.append((old_code, new_code))
            rows.append({
                "kind": "section_header",
                "label": current_section or txt[:40],
                "old_ref": current_section_old,
                "section": current_section,
                "tz": None,
                "norm_segs": list(p["segs"]),
                "expl_segs": [],
                "notes": [],
            })
            continue

        if kind == "number":
            fmts = {s["fmt"] for s in p["segs"] if s["text"].strip()}
            new_digits = plain_text(p["segs"],
                                    fmt_filter={"unchanged", "added", "moved_to"}).strip()
            old_digits = plain_text(p["segs"],
                                    fmt_filter={"unchanged", "deleted", "moved_from"}).strip()

            if fmts == {"deleted"}:
                # Fully deleted old Tz — create a separate "alt" row,
                # keep last_new_tz untouched so the next real Tz still
                # sees the correct expected-next context.
                flush_tz()
                tz_norm.clear(); tz_expl.clear(); tz_notes.clear()
                current_tz = old_digits or txt
                current_tz_old = None
                current_tz_is_alt = True
                continue

            # Normal / added / renumbered case.
            tz_str, _ = split_footnote_from_tz(new_digits or txt, last_new_tz)
            if tz_str is None:
                drop_next_footnote_body = True
                continue
            old_tz_str = None
            if old_digits and old_digits != new_digits and old_digits.isdigit():
                old_tz_str = old_digits
            flush_tz()
            tz_norm.clear(); tz_expl.clear(); tz_notes.clear()
            current_tz = tz_str
            current_tz_old = old_tz_str
            current_tz_is_alt = False
            try:
                last_new_tz = int(tz_str)
            except ValueError:
                pass
            if old_tz_str:
                renames.append((f"{current_section} Tz. {old_tz_str}",
                                f"{current_section} Tz. {tz_str}"))
            continue

        if drop_next_footnote_body and col == "L" and kind == "body":
            drop_next_footnote_body = False
            continue
        drop_next_footnote_body = False

        # Body paragraph → append to current Tz buckets.
        target = tz_norm if col == "L" else tz_expl
        if target:
            target.append({"text": "\n", "fmt": "unchanged",
                           "flags": 0, "size": 10, "font": "Calibri"})
        target.extend(p["segs"])

    flush_tz()
    return rows, renames


_SENT_SPLIT = re.compile(r'(?<=[.;!?])\s+|\n+')


def _ratio(a: str, b: str, cap: int = 900) -> float:
    """Direkter SequenceMatcher-Vergleich, geclipt auf `cap` Zeichen."""
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a[:cap], b[:cap]).ratio()


def _partial_ratio(needle: str, haystack: str, cap: int = 600) -> float:
    """Gleitet `needle` als Fenster über `haystack`; gibt besten ratio zurück.

    Fensterbreite = len(needle) + 25 % Toleranz.
    Ist needle >= haystack, direkter Vergleich.
    """
    if not needle or not haystack:
        return 0.0
    n_clip = needle[:cap]
    if len(needle) >= len(haystack):
        return SequenceMatcher(None, n_clip, haystack[:cap]).ratio()
    window = min(len(needle) + max(len(needle) // 4, 40), len(haystack))
    step = max(1, window // 8)
    best = 0.0
    for start in range(0, len(haystack) - window + 1, step):
        r = SequenceMatcher(None, n_clip,
                            haystack[start:start + window][:cap]).ratio()
        if r > best:
            best = r
    return best


def _sentences(text: str, min_len: int = 55) -> list:
    """Zerlegt Text in Sätze; begrenzt auf 12 Stück."""
    parts = _SENT_SPLIT.split(text)
    return [p.strip() for p in parts if len(p.strip()) >= min_len][:12]


def _position_windows(text: str, min_len: int = 40) -> list:
    """Drei überlappende Fenster (Anfang / Mitte / Ende) mit ~45 % Textlänge.

    Dient dazu, Textteile zu finden, die weder als Absatz noch als ganzer
    Block in einem anderen Tz auftauchen, aber als mittlerer Abschnitt.
    """
    size = max(min_len, int(len(text) * 0.45))
    if size >= len(text):
        return [text]
    centres = [size // 2, len(text) // 2, len(text) - size // 2]
    seen, result = set(), []
    for c in centres:
        start = max(0, c - size // 2)
        w = text[start:start + size].strip()
        if len(w) >= min_len and w not in seen:
            result.append(w)
            seen.add(w)
    return result


def find_tz_moves(rows):
    """5-Pass-Heuristik: gestrichene Tz gegen hinzugefügte matchen.

    Pass 1 – Volltext          : gesamter Textkörper beider Seiten (bis 900 Z.)
    Pass 2 – Absätze vorwärts  : jeder Absatz von dt als Fenster in at
    Pass 3 – Positions­fenster  : Anfang/Mitte/Ende-Block von dt in at
    Pass 4 – Rückwärts         : Absätze und Fenster von at in dt
    Pass 5 – Satz-Fingerabdruck: Anteil gemeinsamer Sätze (≥ 82 % Ähnlichkeit)

    Passes 2–5 werden nur für Kandidaten durchgeführt, die in Pass 1
    mindestens 0,20 Ähnlichkeit erreichen (Vorfilter), um die Laufzeit
    zu begrenzen. Pro gelöschter Tz werden max. 20 Kandidaten tiefgehend
    analysiert; zusätzlich alle mit Pass-1-Score ≥ 0,35.

    Berücksichtigt Normtext und Erläuterung (expl_segs).
    """
    MIN_LEN = 40
    PREFILTER  = 0.20   # Pass-1-Mindest-Score für tiefe Analyse
    TOP_K      = 20     # Anzahl bester Kandidaten für Passes 2–5
    FULL_THR   = 0.55   # Mindest-Score für Volltext-Match (Ausgabe)
    PART_THR   = 0.68   # Mindest-Score für Teiltext-Match (Ausgabe)

    def del_text(r):
        t = plain_text(r["norm_segs"],
                       fmt_filter={"deleted", "moved_from"}).strip()
        e = plain_text(r["expl_segs"],
                       fmt_filter={"deleted", "moved_from"}).strip()
        return (t + ("\n" + e if e else "")).strip()

    def add_text(r):
        t = plain_text(r["norm_segs"],
                       fmt_filter={"added", "moved_to"}).strip()
        e = plain_text(r["expl_segs"],
                       fmt_filter={"added", "moved_to"}).strip()
        return (t + ("\n" + e if e else "")).strip()

    tz_idxs = [i for i, r in enumerate(rows) if r["kind"] == "tz"]
    del_map = {i: del_text(rows[i]) for i in tz_idxs}
    add_map = {i: add_text(rows[i]) for i in tz_idxs}

    del_idxs = [i for i in tz_idxs if len(del_map[i]) >= MIN_LEN]
    add_idxs = [i for i in tz_idxs if len(add_map[i]) >= MIN_LEN]

    # Vorbereitung: Absätze, Positionsfenster, Sätze
    del_paras = {i: [p.strip() for p in del_map[i].split("\n")
                     if len(p.strip()) >= MIN_LEN] for i in del_idxs}
    add_paras = {i: [p.strip() for p in add_map[i].split("\n")
                     if len(p.strip()) >= MIN_LEN] for i in add_idxs}
    del_wins  = {i: _position_windows(del_map[i]) for i in del_idxs}
    add_wins  = {i: _position_windows(add_map[i]) for i in add_idxs}
    del_sents = {i: _sentences(del_map[i]) for i in del_idxs}
    add_sents = {i: _sentences(add_map[i]) for i in add_idxs}

    n = len(del_idxs)
    best_score   = {}
    best_partner = {}
    best_partial = {}

    def register(i, j, score, partial):
        if score > best_score.get(i, 0.0):
            best_score[i]   = score
            best_partner[i] = j
            best_partial[i] = partial

    for idx, i in enumerate(del_idxs, 1):
        if idx % 25 == 0 or idx == 1:
            print(f"  Verschiebungsanalyse {idx}/{n} …")
        dt = del_map[i]

        # Pass 1: Volltext-Vorfilter für alle add-Texte
        p1_scores = []
        for j in add_idxs:
            if j == i:
                continue
            at = add_map[j]
            lr = len(dt) / len(at) if at else 0
            if 0.10 <= lr <= 10.0:
                r = _ratio(dt, at, 900)
                p1_scores.append((r, j))
                register(i, j, r, False)

        # Kandidaten für Passes 2–5: Top-K + alle mit Score ≥ PREFILTER
        p1_scores.sort(reverse=True)
        deep_js = {j for _, j in p1_scores[:TOP_K]}
        deep_js |= {j for r, j in p1_scores if r >= PREFILTER}

        for j in deep_js:
            at = add_map[j]

            # Pass 2: Absätze von dt als Fenster in at
            for para in del_paras[i]:
                r = _partial_ratio(para, at)
                if r >= PART_THR:
                    register(i, j, r, True)

            # Pass 3: Positions­fenster (Anfang/Mitte/Ende) von dt in at
            for win in del_wins[i]:
                r = _partial_ratio(win, at)
                if r >= PART_THR:
                    register(i, j, r, True)

            # Pass 4: Rückwärts — Absätze und Fenster von at in dt
            for para in add_paras[j]:
                r = _partial_ratio(para, dt)
                if r >= PART_THR:
                    register(i, j, r, True)
            for win in add_wins[j]:
                r = _partial_ratio(win, dt)
                if r >= PART_THR:
                    register(i, j, r, True)

            # Pass 5: Satz-Fingerabdruck
            ds_list = del_sents[i]
            as_list = add_sents[j]
            if ds_list and as_list:
                hits = sum(
                    1 for d in ds_list
                    if any(_ratio(d, a, 300) >= 0.82 for a in as_list)
                )
                if hits:
                    register(i, j, hits / len(ds_list), hits < len(ds_list))

    # Ergebnisse eintragen
    for i in del_idxs:
        score = best_score.get(i, 0.0)
        j     = best_partner.get(i)
        if j is None or score < FULL_THR:
            continue
        partial    = best_partial.get(i, False)
        kind_label = "Teilverschiebung" if partial else "Verschiebung"
        rows[i].setdefault("G", rows[j]["label"])
        rows[i]["notes"].append(
            f"mögliche {kind_label} nach {rows[j]['label']} "
            f"(Ähnlichkeit {score:.0%})"
        )
        rows[j].setdefault("G", rows[i]["label"])
        rows[j]["notes"].append(
            f"mögliche Herkunft aus {rows[i]['label']} "
            f"(Ähnlichkeit {score:.0%})"
        )
        if score < 0.75:
            rows[i]["uncertain"] = True
            rows[j]["uncertain"] = True


def write_excel(rows, renames=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Änderungen pro Tz"
    headers = ["Textziffer", "alte Referenz", "Normtext", "Erläuterung",
               "Änderungsart Normtext", "Änderungsart Erläuterung",
               "Verschiebung", "Unsicher", "Anmerkungen"]
    ws.append(headers)
    header_font = Font(bold=True)
    for c in ws[1]:
        c.font = header_font
        c.alignment = Alignment(horizontal="left", vertical="top")

    fills = {
        "unverändert":       PatternFill("solid", fgColor="FFFFFFFF"),
        "geändert":          PatternFill("solid", fgColor="FFDEEBF7"),
        "gestrichen":        PatternFill("solid", fgColor="FFFCE4E4"),
        "hinzugefügt":       PatternFill("solid", fgColor="FFE2EFDA"),
        "verschoben vorher": PatternFill("solid", fgColor="FFFFF2CC"),
        "verschoben nachher":PatternFill("solid", fgColor="FFEDFADE"),
    }
    heading_fill = PatternFill("solid", fgColor="FFEDE7F6")
    # For the overall row colour use the "strongest" change of the two
    # columns — geändert > gestrichen/hinzugefügt > verschoben > unverändert.
    severity = {
        "unverändert": 0, "verschoben nachher": 1, "verschoben vorher": 1,
        "hinzugefügt": 2, "gestrichen": 2, "geändert": 3,
    }

    for r in rows:
        norm_rich = segs_to_rich(r["norm_segs"])
        expl_rich = segs_to_rich(r["expl_segs"])
        d_norm = classify_entry(r["norm_segs"]) if r["norm_segs"] else "unverändert"
        d_expl = classify_entry(r["expl_segs"]) if r["expl_segs"] else "unverändert"
        g = r.get("G", "")
        uncertain = "Ja" if r.get("uncertain") else "Nein"
        auto_notes = list(r.get("notes", []))
        norm_sum = diff_summary(r["norm_segs"])
        if norm_sum:
            auto_notes.append(f"Normtext: {norm_sum}")
        expl_sum = diff_summary(r["expl_segs"])
        if expl_sum:
            auto_notes.append(f"Erläuterung: {expl_sum}")
        notes = "; ".join(auto_notes)

        old_ref = r.get("old_ref", "")
        if r["kind"] == "section_header":
            heading_note = notes or "Überschrift"
            ws.append([r["label"], old_ref, norm_rich, "",
                       d_norm, "", "", "", heading_note])
            row_idx = ws.max_row
            fill = heading_fill if d_norm != "unverändert" else fills["unverändert"]
        else:
            ws.append([r["label"], old_ref, norm_rich, expl_rich,
                       d_norm, d_expl, g, uncertain, notes])
            row_idx = ws.max_row
            worst = max(severity.get(d_norm, 0), severity.get(d_expl, 0))
            if worst == 0:
                d_row = "unverändert"
            elif worst == 3:
                d_row = "geändert"
            else:
                # pick the first non-unverändert
                d_row = d_norm if severity.get(d_norm, 0) >= severity.get(d_expl, 0) else d_expl
            fill = fills.get(d_row, fills["unverändert"])

        for col_i in range(1, 10):
            cell = ws.cell(row=row_idx, column=col_i)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = {"A": 20, "B": 20, "C": 70, "D": 70, "E": 18, "F": 18,
              "G": 18, "H": 10, "I": 38}
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w
    ws.freeze_panes = "A2"

    # Legende
    leg = wb.create_sheet("Legende")
    leg.append(["Spalte / Kennzeichnung", "Bedeutung"])
    leg["A1"].font = header_font
    leg["B1"].font = header_font
    for a, b in [
        ("A – Textziffer",            "neuer Abschnitt + Tz-Nummer, z.B. 'AT 1 Tz. 3'"),
        ("B – alte Referenz",         "alter Pfad der Tz vor Umbenennung/Umnummerierung; leer wenn unverändert. Beispiel: 'AT 4.4.2 Tz. 6' für die heutige 'AT 4.4.2 Tz. 5'"),
        ("C – Normtext",              "Rich-Text des gesamten linken Spaltentexts der Tz"),
        ("D – Erläuterung",           "Rich-Text aller rechten Spalten-Absätze, die zu der Tz gehören"),
        ("E – Änderungsart Normtext",      "unverändert / geändert / gestrichen / hinzugefügt / verschoben"),
        ("F – Änderungsart Erläuterung",   "dasselbe Schema für die Erläuterung"),
        ("G – Verschiebung",          "Heuristische Ziel-/Herkunfts-Tz für inhaltliche Verschiebungen (Volltext oder Teiltext)"),
        ("H – Unsicher",              "'Ja' = Verschiebungs-Match unter 75 % Ähnlichkeit"),
        ("I – Anmerkungen",           "Diff-Summary (Wortzahlen, Umformulierungs-Hinweis) und Verschiebungs-Vermerke mit Ähnlichkeit in %. 'Teilverschiebung' = nur Teile des Textkörpers erscheinen an anderer Stelle."),
        ("", ""),
        ("schwarzer Text in C/D",                   "unverändert"),
        ("rot + Durchstreichung",                    "gestrichen (alt)"),
        ("rot + Unterstreichung",                    "neu hinzugefügt"),
        ("grün + Durchstreichung",                   "verschoben (alte Stelle)"),
        ("grün + Unterstreichung",                   "verschoben (neue Stelle)"),
        ("Zeilenhintergrund weiß",                   "Eintrag unverändert"),
        ("Zeilenhintergrund hellblau (#DEEBF7)",     "Tz insgesamt geändert"),
        ("Zeilenhintergrund hellrot (#FCE4E4)",      "Tz gestrichen"),
        ("Zeilenhintergrund hellgrün (#E2EFDA)",     "Tz neu hinzugefügt"),
        ("Zeilenhintergrund hellgelb (#FFF2CC)",     "Verschoben vorher"),
        ("Zeilenhintergrund gelbgrün (#EDFADE)",     "Verschoben nachher"),
        ("Zeilenhintergrund helllila (#EDE7F6)",     "Überschrift geändert"),
    ]:
        leg.append([a, b])
    leg.column_dimensions["A"].width = 44
    leg.column_dimensions["B"].width = 60
    for row in leg.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")

    # Umbenennungen sheet — quick lookup alt → neu.
    if renames:
        un = wb.create_sheet("Umbenennungen")
        un.append(["alter Code", "neuer Code"])
        un["A1"].font = header_font
        un["B1"].font = header_font
        for old, new in renames:
            un.append([old, new])
        un.column_dimensions["A"].width = 18
        un.column_dimensions["B"].width = 18
        un.freeze_panes = "A2"
        for row in un.iter_rows(min_row=2):
            for c in row:
                c.alignment = Alignment(vertical="top")

    wb.save(OUT)


def main():
    doc = fitz.open(PDF)
    all_segs = []
    print(f"Parsing {len(doc)} pages…")
    for pno in range(len(doc)):
        page = doc[pno]
        segs = load_page(page)
        for s in segs:
            s["page"] = pno
        all_segs.extend(segs)
        if (pno + 1) % 20 == 0:
            print(f"  … page {pno+1}/{len(doc)}")

    paragraphs = group_paragraphs(all_segs)

    # Drop everything before the first real section heading.
    first_section = next((k for k, p in enumerate(paragraphs)
                          if p["kind"] == "section"
                          and SECTION_RE.match(plain_text(p["segs"]).strip() or "")),
                         0)
    paragraphs = paragraphs[first_section:]

    # Reading-order per page (y, then col) so Erläuterungen land at the
    # correct Tz.
    by_page = {}
    for p in paragraphs:
        by_page.setdefault(p["page"], []).append(p)
    ordered = []
    for pno in sorted(by_page):
        ordered.extend(sorted(by_page[pno],
                              key=lambda p: (round(p["y_top"], 0),
                                             0 if p["col"] == "L" else 1)))
    paragraphs = ordered

    rows, renames = build_tz_rows(paragraphs)
    print(f"Got {len(rows)} rows "
          f"({sum(1 for r in rows if r['kind']=='section_header')} Überschriften, "
          f"{sum(1 for r in rows if r['kind']=='tz')} Textziffern, "
          f"{len(renames)} Umbenennungen)")

    find_tz_moves(rows)

    write_excel(rows, renames)
    _postprocess_xlsx(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
